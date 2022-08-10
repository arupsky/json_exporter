from lib2to3.pgen2.token import NEWLINE
from os import listdir
import openpyxl
import Trial_DTO
import pandas

# input_folder_path = 'test_data/118170'+'/'
input_folder_path = 'source_data/119509'+'/'
output_folder_path = input_folder_path +'/output.json'
excel_file_loc = ''
csv_file_loc = ''
for file in listdir(input_folder_path):
    if file.endswith('.csv'):
        csv_file_loc = input_folder_path + file
    elif file.endswith('.xlsx'):
        excel_file_loc = input_folder_path + file


trial_dictionary = {}

trial_id = ""
pupil_dilation = []
baseline = []
rating = 0

wb_obj =  openpyxl.load_workbook(excel_file_loc)
current_row = 2
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row

isSaved = False
trial_DTO = Trial_DTO.Trial_DTO()
# trial_DTO = Trial_DTO.build_trial_dto()
for current_row in range(2, m_row + 1):
    diameter_obj = sheet_obj.cell(row = current_row, column = 1)
    stimulii_obj = sheet_obj.cell(row = current_row, column = 2)
    # print(diameter_obj.value)
    if (stimulii_obj.value == 'b'):
        baseline.append(diameter_obj.value)
        isSaved = False
    elif (stimulii_obj.value == 'r'):
        if not isSaved:
            # save to dictionary
            trial_DTO = Trial_DTO.Trial_DTO(trial_id, pupil_dilation, baseline, 0)        
            # print(str(trial_DTO.baseline))

            trial_dictionary[trial_id] = trial_DTO

            pupil_dilation = []
            baseline = []
            isSaved = True
            
    else:        
        pupil_dilation.append(diameter_obj.value)
        trial_id = stimulii_obj.value


rating_csv = pandas.read_csv(csv_file_loc, usecols=['image.dir', 'rating_score.response'])

for index in range (0, len(rating_csv)):
    trial_dictionary[rating_csv['image.dir'][index]].rating = rating_csv['rating_score.response'][index]

# some weird fucking reason json serialization / marshal could not be used, manual implemenation is done instead

serialized_json_string = '{"trials":['

for key in trial_dictionary:
    serialized_json_string += '\n' + str (trial_dictionary[key].tojson()) + ','

serialized_json_string = serialized_json_string[:-1]
serialized_json_string += ']}'

f = open (output_folder_path, 'w')
f.write (serialized_json_string)
f.close()

print ("Done: " + str (len(rating_csv)))